# format_ortho_signals.py
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

# ---- paths ----
in_csv  = "../results/ortho_signals_grouped.csv"
out_xlsx = "../results/ortho_signals_grouped_formatted.xlsx"

# ---- load ----
df = pd.read_csv(in_csv)

# Ensure expected columns exist
required = {"Ortho Program", "Rank", "Region", "OrthoCultureScoreUsed"}
missing = required - set(df.columns)
if missing:
    raise ValueError(f"Missing columns in CSV: {missing}")

# ---- region color map (edit as you like) ----
region_colors = {
    "Boston": "#0096C7",           # blue
    "LA": "#ED80E9",               # pink
    "San Francisco": "#9ACD32",    # light green  (fixed double #)
    "NYC area": "#CCCCCC",         # grey
    "Pennsylvania": "#9370DB",     # purple-ish
    "Texas": "#FFA500",            # orange
    "South (not NC)": "#FFD700",   # gold
    "NC": "#87CEEB",               # sky blue
    "Chicago / Michigan": "#ED2939",
    "Midwest": "#1E90FF",
    "West": "#CCAA14",
    "Other": "#E4A0F7",
    "Unknown": "#FFFFFF",
}

# ---- write to workbook ----
wb = Workbook()
ws = wb.active
ws.title = "Ortho Signals"

# headers
ws.append(list(df.columns))

# rows
for _, row in df.iterrows():
    ws.append(list(row.values))

# locate columns
headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
col_prog   = headers["Ortho Program"]
col_rank   = headers["Rank"]
col_cult   = headers["OrthoCultureScoreUsed"]
col_region = headers["Region"]

n_rows = ws.max_row

# ---- conditional formatting: Rank (low=green, high=red) ----
rank_col_letter = get_column_letter(col_rank)
ws.conditional_formatting.add(
    f"{rank_col_letter}2:{rank_col_letter}{n_rows}",
    ColorScaleRule(
        start_type="min", start_color="FF63BE7B",  # green
        end_type="max",   end_color="FFFF0000"     # red
    )
)

# ---- conditional formatting: OrthoCultureScoreUsed (red -> white @ 0 -> green) ----
# Compute min/max for the scale
cult_min = float(df["OrthoCultureScoreUsed"].min())
cult_max = float(df["OrthoCultureScoreUsed"].max())

cult_col_letter = get_column_letter(col_cult)
ws.conditional_formatting.add(
    f"{cult_col_letter}2:{cult_col_letter}{n_rows}",
    ColorScaleRule(
        start_type="num", start_value=cult_min, start_color="FFFF0000",  # red
        mid_type="num",   mid_value=0,        mid_color="FFFFFFFF",      # white at 0
        end_type="num",   end_value=cult_max, end_color="FF63BE7B"       # green
    )
)

# ---- region-based highlighting on BOTH the Ortho Program and Region cells ----
for r in range(2, n_rows + 1):
    region_val = ws.cell(row=r, column=col_region).value
    hex_color = region_colors.get(region_val, "#FFFFFF").replace("#", "")
    fill = PatternFill(fill_type="solid", fgColor=hex_color)
    ws.cell(row=r, column=col_prog).fill = fill
    ws.cell(row=r, column=col_region).fill = fill  # <-- also color the Region cell

# (optional) widen columns a bit
for col_idx in range(1, ws.max_column + 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 32

wb.save(out_xlsx)
print(f"Saved: {out_xlsx}")